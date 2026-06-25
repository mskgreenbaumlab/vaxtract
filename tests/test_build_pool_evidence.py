"""Deterministic per-patient pool-immunogenicity evidence (`build_pool_evidence`, A' design).

33064988's per-patient pool-ELISpot evidence (Fig 4A / Supp Fig 3) is stated uniformly in text
("responses in all patients") but only figure-quantified -> the agent built it by hand and the
count swung 0<->34 run-to-run. build_pool_evidence emits one pool/immunogenic row per MONITORED
patient deterministically, magnitude=null + needs_review, figure-provenance. Faithfulness: only
patients with an actual pool (= individually shown) get a row.
"""
import json
import pathlib

import openpyxl

import agent_core

PKT = pathlib.Path(__file__).resolve().parents[1]
REF = json.loads((PKT / "reference_records" / "rojas_extracted.json").read_text())
META = {k: v for k, v in REF.items() if not isinstance(v, list)}


def _partial(out):
    return json.loads(pathlib.Path(str(out) + ".partial.json").read_text())


def _init_with_pools(out, patients):
    """Partial record with one minimal ExtractedPeptidePool per patient in `patients`."""
    pathlib.Path(out).parent.mkdir(parents=True, exist_ok=True)
    ok, msg = agent_core.init_partial(str(out), json.dumps(META))
    assert ok, msg
    pools = [{"paper_local_id": f"POOL-{p}", "patient_paper_id": p,
              "member_peptide_ids": [f"IMP-{p}-1"], "quoted_text": f"{p} pool",
              "section_ref": "Vaccine peptides"} for p in patients]
    ok, msg = agent_core.append_section(str(out), "pools", json.dumps(pools))
    assert ok, msg


def test_one_immunogenic_row_per_monitored_patient(tmp_path):
    out = tmp_path / "rec.json"
    _init_with_pools(out, ["B1", "M2", "L3"])
    ok, msg = agent_core.build_pool_evidence(str(out), patients=["B1", "M2", "L3"])
    assert ok, msg
    ev = _partial(out)["evidence"]
    assert len(ev) == 3
    for e in ev:
        assert e["target_kind"] == "pool"
        assert e["pool_paper_id"] == f"POOL-{e['patient_paper_id']}"
        assert e["outcome"] == "immunogenic"
        assert e["assay"] == "elispot"
        assert e["vaccine_induced"] is True
        assert e["magnitude"] is None
        assert e["needs_review"] is True
        # provenance is figure-derived (distinct from readable-table rows)
        assert e["provenance"][0]["kind"] == "figure"


def test_monitored_set_derived_from_sheet_names(tmp_path):
    out = tmp_path / "rec.json"
    _init_with_pools(out, ["B1", "M13"])
    xlsx = tmp_path / "mmc2.xlsx"
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    wb.create_sheet("Vaccine peptides")
    wb.create_sheet("IAP-B1"); wb.create_sheet("IAP-M13")
    wb.save(xlsx)
    ok, msg = agent_core.build_pool_evidence(str(out), sheets_path=str(xlsx), sheet_pattern="IAP-(.+)")
    assert ok, msg
    pts = sorted(e["patient_paper_id"] for e in _partial(out)["evidence"])
    assert pts == ["B1", "M13"]


def test_does_not_invent_rows_for_patients_without_a_pool(tmp_path):
    # FAITHFULNESS: M99 is "monitored" but has no pool -> no row (the abstract's "all patients"
    # must not manufacture a per-patient finding the paper never shows individually).
    out = tmp_path / "rec.json"
    _init_with_pools(out, ["B1", "M2"])           # pools only for B1, M2
    ok, msg = agent_core.build_pool_evidence(str(out), patients=["B1", "M2", "M99"])
    assert ok, msg
    pts = sorted(e["patient_paper_id"] for e in _partial(out)["evidence"])
    assert pts == ["B1", "M2"]
    assert "M99" in msg  # reported as skipped


def test_deterministic_across_two_calls(tmp_path):
    seen = []
    for i in range(2):
        out = tmp_path / f"r{i}" / "rec.json"
        _init_with_pools(out, ["B1", "M2"])
        agent_core.build_pool_evidence(str(out), patients=["B1", "M2"])
        ev = _partial(out)["evidence"]
        seen.append(sorted((e["patient_paper_id"], e["pool_paper_id"], e["outcome"]) for e in ev))
    assert seen[0] == seen[1] == [("B1", "POOL-B1", "immunogenic"), ("M2", "POOL-M2", "immunogenic")]


def test_errors_without_pools(tmp_path):
    out = tmp_path / "rec.json"
    pathlib.Path(out).parent.mkdir(parents=True, exist_ok=True)
    agent_core.init_partial(str(out), json.dumps(META))   # no pools
    ok, msg = agent_core.build_pool_evidence(str(out), patients=["B1"])
    assert not ok
    assert "build_pools" in msg


def test_errors_without_a_monitored_set(tmp_path):
    out = tmp_path / "rec.json"
    _init_with_pools(out, ["B1"])
    ok, msg = agent_core.build_pool_evidence(str(out))  # no patients, no sheets
    assert not ok
    assert "patients" in msg or "sheet" in msg
