"""read_table_rows oversized-sheet + underline fidelity (root-cause fixes from the
Keskin + Li new-paper test runs, 2026-06-04).

Keskin Table S5 (the 103-peptide master) was UNREADABLE: a naive dump of thousands of
rows with multi-kB neoORF-context cells blew the CLI's ~64KB per-result cap and spilled
to a file the agent may not read. Li's 42 minimal epitopes were invisible: they are
UNDERLINED substrings inside the neoantigen long-peptides, and a plain-text read strips
the formatting. These tests pin: per-cell clipping + byte-cap + true-total reporting,
title-row auto-skip for column projection, and <u>…</u> underline recovery (with runs
merged across a coloured mutant residue).
"""
import json
import pathlib

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

import agent_core

PKT = pathlib.Path(__file__).resolve().parents[1]


# ---- Fix #1: oversized sheet never spills, true total still reported -------------

def test_oversized_sheet_is_byte_capped_with_true_total(tmp_path):
    x = tmp_path / "big.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["LID", "CONTEXT"])
    big = "M" * 5000  # a multi-kB neoORF-context cell, like Keskin S5
    for i in range(400):
        ws.append([f"P{i}", big])
    wb.save(x)

    out = agent_core.read_table_rows(str(x))           # raw preview
    assert len(out.encode()) <= agent_core._MAX_RESULT_BYTES + 2000  # bounded, never 64KB+
    assert "byte-capped" in out
    # the giant cell is clipped, not dumped whole
    assert "M" * 5000 not in out
    assert "+4760 chars" in out  # 5000 - _MAX_CELL_CHARS(240)


def test_oversized_projection_reports_full_total(tmp_path):
    x = tmp_path / "big.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["LID", "CTX"])
    for i in range(400):
        ws.append([f"P{i}", "Q" * 3000])
    wb.save(x)
    # projecting the GIANT column keeps the result big -> cap engages, true total stays
    out = agent_core.read_table_rows(str(x), columns=["LID", "CTX"])
    assert "400/400 data rows" in out          # TRUE total surfaced even when capped
    assert "byte-capped" in out
    body = json.loads(out.split(":\n", 1)[1])
    assert len(body) < 400 and body[0]["LID"] == "P0"  # capped slice, in order

    # projecting AWAY the megacell makes it fit -> all 400 returned, uncapped
    out2 = agent_core.read_table_rows(str(x), columns=["LID"])
    assert "400/400 data rows" in out2 and "byte-capped" not in out2
    assert len(json.loads(out2.split(":\n", 1)[1])) == 400


# ---- Fix #1b: leading title row is auto-skipped for column projection ------------

def test_title_row_autoskipped_so_columns_resolve(tmp_path):
    x = tmp_path / "titled.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Table S2. Selective neoantigens", None, None])  # title row first
    ws.append(["Mutation", "Seq", "ELISPOT"])                    # real headers
    ws.append(["Tmem101.G96V", "QLASTYTAYIV", "-"])
    ws.append(["Lrrc27.G330A", "FKGILPNLPSA", "++++"])
    wb.save(x)
    out = agent_core.read_table_rows(str(x), columns=["Mutation", "ELISPOT"])
    rows = json.loads(out.split(":\n", 1)[1])
    assert rows == [{"Mutation": "Tmem101.G96V", "ELISPOT": "-"},
                    {"Mutation": "Lrrc27.G330A", "ELISPOT": "++++"}]


def test_header_row_can_be_forced(tmp_path):
    x = tmp_path / "titled.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["junk", None]); ws.append(["GENE", "VAL"]); ws.append(["TP53", 7])
    wb.save(x)
    out = agent_core.read_table_rows(str(x), columns=["GENE"], header_row=1)
    assert json.loads(out.split(":\n", 1)[1]) == [{"GENE": "TP53"}]


# ---- Fix #2: underlined minimal-epitope substrings are recoverable --------------

def _rich(prefix, under, suffix, *, split_under=None):
    blocks = [TextBlock(InlineFont(), prefix)]
    if split_under:  # underline interrupted by a (still-underlined) coloured residue
        a, b = split_under
        blocks += [TextBlock(InlineFont(u="single"), a),
                   TextBlock(InlineFont(u="single", color="FF0000"), b)]
    else:
        blocks.append(TextBlock(InlineFont(u="single"), under))
    blocks.append(TextBlock(InlineFont(), suffix))
    return CellRichText(blocks)


def test_underline_wraps_minimal_epitope(tmp_path):
    x = tmp_path / "u.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Mutation", "Seq"])
    ws["B2"] = _rich("QLA", "STYTAYIV", "GYVHYGDWLK")
    ws["A2"] = "Tmem101.G96V"
    wb.save(x)
    out = agent_core.read_table_rows(str(x), columns=["Seq"], underline=True)
    rows = json.loads(out.split(":\n", 1)[1])
    assert rows == [{"Seq": "QLA<u>STYTAYIV</u>GYVHYGDWLK"}]


def test_underline_merges_runs_split_by_mutant_residue(tmp_path):
    # Li Tmem101: underline is two runs (STYTAYI + V) split by a coloured residue.
    x = tmp_path / "u2.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Seq"])
    ws["A2"] = _rich("QLA", None, "GYV", split_under=("STYTAYI", "V"))
    wb.save(x)
    out = agent_core.read_table_rows(str(x), underline=True)
    assert "QLA<u>STYTAYIV</u>GYV" in out          # one contiguous <u>…</u>, not two


def test_underline_off_by_default_leaves_plain_text(tmp_path):
    x = tmp_path / "u3.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Seq"]); ws["A2"] = _rich("QLA", "STYTAYIV", "GYV")
    wb.save(x)
    out = agent_core.read_table_rows(str(x))         # underline defaults False
    assert "<u>" not in out and "QLASTYTAYIVGYV" in out
