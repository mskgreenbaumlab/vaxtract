import json
import pathlib
import openpyxl
import agent_core

PKT = pathlib.Path(__file__).resolve().parents[1]
REF = json.loads((PKT / "reference_records" / "rojas_extracted.json").read_text())
META = {k: v for k, v in REF.items() if not isinstance(v, list)}
IMPS = REF["immunizing_peptides"][:2]


def _write_sheet(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["LID", "SEQ", "GENE", "RESP"])
    for p in IMPS:
        ws.append([p["paper_local_id"], p["sequence"], p.get("gene_symbol", ""), "De novo response"])
    ws.append(["X9", "", "ZZZ", "no"])  # junk row (empty sequence) for filter/atomic tests
    wb.save(path)


PEPTIDE_FIELDS = {
    "paper_local_id": {"col": "LID"},
    "sequence": {"col": "SEQ"},
    "gene_symbol": {"col": "GENE"},
    "is_neoantigen": {"const": True},
    "quoted_text": {"const": "neoantigen target listed in Supp Table 5"},
    "section_ref": {"const": "Supplementary Table 5"},
}


def test_read_xlsx_dicts(tmp_path):
    x = tmp_path / "t.xlsx"; _write_sheet(x)
    headers, rows = agent_core._read_xlsx_dicts(str(x))
    assert headers == ["LID", "SEQ", "GENE", "RESP"]
    assert rows[0]["SEQ"] == IMPS[0]["sequence"]


def test_table_to_entities_appends_valid_rows(tmp_path):
    out = tmp_path / "r.json"; x = tmp_path / "t.xlsx"; _write_sheet(x)
    agent_core.init_partial(str(out), json.dumps(META))
    mapping = {"filter": {"col": "RESP", "in": ["De novo response"]}, "fields": PEPTIDE_FIELDS}
    ok, msg = agent_core.table_to_entities(str(out), "immunizing_peptides", str(x), json.dumps(mapping))
    assert ok, msg
    part = json.loads((tmp_path / "r.json.partial.json").read_text())
    assert len(part["immunizing_peptides"]) == 2          # junk row filtered out
    assert part["immunizing_peptides"][0]["sequence"] == IMPS[0]["sequence"]


def test_table_to_entities_unknown_column_reports(tmp_path):
    out = tmp_path / "r.json"; x = tmp_path / "t.xlsx"; _write_sheet(x)
    agent_core.init_partial(str(out), json.dumps(META))
    mapping = {"fields": {"sequence": {"col": "NOPE"}}}
    ok, msg = agent_core.table_to_entities(str(out), "immunizing_peptides", str(x), json.dumps(mapping))
    assert not ok and "not found" in msg.lower()


def test_table_to_entities_atomic_on_invalid_row(tmp_path):
    out = tmp_path / "r.json"; x = tmp_path / "t.xlsx"; _write_sheet(x)
    agent_core.init_partial(str(out), json.dumps(META))
    # no filter -> junk row (empty sequence) included -> required `sequence` fails -> append nothing
    mapping = {"fields": PEPTIDE_FIELDS}
    ok, msg = agent_core.table_to_entities(str(out), "immunizing_peptides", str(x), json.dumps(mapping))
    assert not ok
    part = json.loads((tmp_path / "r.json.partial.json").read_text())
    assert part["immunizing_peptides"] == []              # atomic: nothing appended


def test_table_to_entities_requires_partial(tmp_path):
    x = tmp_path / "t.xlsx"; _write_sheet(x)
    mapping = {"fields": PEPTIDE_FIELDS}
    ok, msg = agent_core.table_to_entities(str(tmp_path / "none.json"), "immunizing_peptides", str(x), json.dumps(mapping))
    assert not ok and "init_record" in msg


def test_table_to_entities_rejects_non_dict_rule(tmp_path):
    out = tmp_path / "r.json"; x = tmp_path / "t.xlsx"; _write_sheet(x)
    agent_core.init_partial(str(out), json.dumps(META))
    mapping = {"fields": {"sequence": "col_name"}}  # rule is a string, not an object
    ok, msg = agent_core.table_to_entities(str(out), "immunizing_peptides", str(x), json.dumps(mapping))
    assert not ok and "object" in msg.lower()       # reported, not raised


def test_table_to_entities_filter_missing_column_reported(tmp_path):
    out = tmp_path / "r.json"; x = tmp_path / "t.xlsx"; _write_sheet(x)
    agent_core.init_partial(str(out), json.dumps(META))
    mapping = {"filter": {"col": "NOPE", "in": ["x"]}, "fields": PEPTIDE_FIELDS}
    ok, msg = agent_core.table_to_entities(str(out), "immunizing_peptides", str(x), json.dumps(mapping))
    assert not ok and "not found" in msg.lower()


def test_table_to_entities_empty_sheet_reported(tmp_path):
    import openpyxl
    x = tmp_path / "empty.xlsx"; openpyxl.Workbook().save(x)   # a sheet with no rows
    out = tmp_path / "r.json"; agent_core.init_partial(str(out), json.dumps(META))
    ok, msg = agent_core.table_to_entities(str(out), "immunizing_peptides", str(x), json.dumps({"fields": PEPTIDE_FIELDS}))
    assert not ok and "no rows" in msg.lower()       # readable, not an opaque error


# ---- v2.11.3 root-cause fix #2: multi-sheet add_table + `extract` (one mapping over per-patient sheets) ----

def _write_multisheet(path):
    """Mimic 33064988: one immunogenicity sheet per patient (IAP-<patient>), shared columns."""
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for sh, seqs in (("IAP-M1", ["SLLQHLIGL", "SIINFEKL"]), ("IAP-M13", ["KVAELVHFL"])):
        ws = wb.create_sheet(sh)
        ws.append(["PEP", "SEQ"])
        for i, s in enumerate(seqs):
            ws.append([f"{sh}-p{i}", s])
    wb.save(path)


_MS_FIELDS = {
    "paper_local_id": {"col": "PEP"},
    "sequence": {"col": "SEQ"},
    "is_neoantigen": {"const": True},
    "patient_paper_id": {"col": "__sheet__", "extract": "IAP-(.+)"},  # 'IAP-M13' -> 'M13'
    "quoted_text": {"const": "per-patient immunogenicity sheet"},
    "section_ref": {"const": "Supp IAP"},
}


def test_multisheet_add_table_loads_all_sheets_one_call(tmp_path):
    out = tmp_path / "r.json"; x = tmp_path / "ms.xlsx"; _write_multisheet(x)
    agent_core.init_partial(str(out), json.dumps(META))
    ok, msg = agent_core.table_to_entities(str(out), "immunizing_peptides", str(x),
                                           json.dumps({"fields": _MS_FIELDS}),
                                           sheets=["IAP-M1", "IAP-M13"])
    assert ok, msg
    peps = json.loads((tmp_path / "r.json.partial.json").read_text())["immunizing_peptides"]
    assert len(peps) == 3                                   # 2 + 1 across both sheets, one call
    assert [p["patient_paper_id"] for p in peps] == ["M1", "M1", "M13"]   # __sheet__ + extract
    assert "from 2 sheets" in msg


def test_multisheet_atomic_across_sheets(tmp_path):
    # a bad row in the SECOND sheet -> nothing from EITHER sheet is appended
    out = tmp_path / "r.json"; x = tmp_path / "ms.xlsx"
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    ws1 = wb.create_sheet("IAP-M1"); ws1.append(["PEP", "SEQ"]); ws1.append(["a", "SLLQHLIGL"])
    ws2 = wb.create_sheet("IAP-M2"); ws2.append(["PEP", "SEQ"]); ws2.append(["b", "not-a-sequence!"])
    wb.save(x)
    agent_core.init_partial(str(out), json.dumps(META))
    ok, _ = agent_core.table_to_entities(str(out), "immunizing_peptides", str(x),
                                         json.dumps({"fields": _MS_FIELDS}), sheets=["IAP-M1", "IAP-M2"])
    assert not ok
    assert json.loads((tmp_path / "r.json.partial.json").read_text())["immunizing_peptides"] == []


def test_extract_regex_in_mapping():
    import table_map
    row = {"__sheet__": "IAP-M13", "SEQ": "SLLQHLIGL"}
    out = table_map.apply_mapping(row, {"patient_paper_id": {"col": "__sheet__", "extract": "IAP-(.+)"},
                                        "sequence": {"col": "SEQ"}})
    assert out["patient_paper_id"] == "M13" and out["sequence"] == "SLLQHLIGL"
    # no match -> value unchanged (safe fallback)
    out2 = table_map.apply_mapping({"x": "plain"}, {"f": {"col": "x", "extract": "ZZZ-(.+)"}})
    assert out2["f"] == "plain"


def test_add_table_tool_logs_its_mode():
    # observability (2026-06-09): the add_table MCP wrapper must emit a greppable '[add_table]' line
    # recording single vs multi(N), so the per-paper log shows whether the bulk sheets=[...] path was
    # used. Guard the marker so a refactor can't silently drop it (we grep logs for it post-run).
    src = (PKT / "vaxtract" / "extraction_agent.py").read_text()
    body = src[src.index("async def add_table"):src.index("server = create_sdk_mcp_server")]
    assert "[add_table]" in body and "multi(" in body and "mode=" in body
